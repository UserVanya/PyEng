import os
import tkinter as tk
from tkinter.messagebox import*
import openpyxl
import json
import pandas as pd
import ctypes
import sys
from deep_translator import GoogleTranslator
from pyeng_yacl_translator_impl import YacloudTranslator
from deep_translator import PonsTranslator
def is_ru_lang_keyboard():
    '''
    Auxiliary function to check if the current keyboard layout is Russian.
    It is needed to handle Ctrl+C, Ctrl+V, Ctrl+X, Ctrl+A, Delete keys because if the keyboard layout is Russian,
    the Ctrl+C, Ctrl+V, Ctrl+X, Ctrl+A, Delete keys cannot be handled by the tkinter library.
    '''
    u = ctypes.windll.LoadLibrary("user32.dll")
    pf = getattr(u, "GetKeyboardLayout")
    return hex(pf(0)) == '0x4190419'
def ru_keys_handler(event):
    '''
    Auxiliary function to check if the current keyboard layout is Russian.
    It is needed to handle Ctrl+C, Ctrl+V, Ctrl+X, Ctrl+A, Delete keys because if the keyboard layout is Russian,
    the Ctrl+C, Ctrl+V, Ctrl+X, Ctrl+A, Delete keys cannot be handled by the tkinter library.
    '''
    if is_ru_lang_keyboard():
        if event.keycode==86:
            event.widget.event_generate("<<Paste>>")
        if event.keycode==67: 
            event.widget.event_generate("<<Copy>>")    
        if event.keycode==88: 
            event.widget.event_generate("<<Cut>>")    
        if event.keycode==65535: 
            event.widget.event_generate("<<Clear>>")
        if event.keycode==65: 
            event.widget.event_generate("<<SelectAll>>")
class PyengCore:
    '''
    A core class for PyEng application, stores data, translates words, saves requests history
    '''
    def __init__(self) -> None:
        self.default_headers = ['Word', 'Translation', 'Hint']
        self.save_to_histoty_limit = 30
        self.__init_settings_file()
        self.__init_dictionaries()
        self.tr_impl = YacloudTranslator(self.settings['api_key'], self.settings['folder_id'])
        self.__init_lang_dicts()
    def __init_lang_dicts(self):
        '''
        Initializes two fields: lang_to_code and code_to_lang which are dictionaries to convert language name to language code and vice versa
        It is required to set field tr_impl before calling this method
        '''
        yacl_langs_list = self.tr_impl.get_available_langs()
        
        google_langs_dict = GoogleTranslator().get_supported_languages(as_dict=True)
        self.lang_to_code = {}
        self.code_to_lang = {}
        for el in yacl_langs_list:
            if 'name' in el.keys() and el['code'].lower() in google_langs_dict.values():
                self.lang_to_code[el['name'].lower()] = el['code'].lower()
                self.code_to_lang[el['code'].lower()] = el['name'].lower()
    def get_lang_code(self, lang):
        '''
        Returns the language code of the language name
        '''
        return self.lang_to_code[lang]
    def get_lang_from_code(self, code):
        '''
        Returns the language name of the language code
        '''
        return self.code_to_lang[code]
    def get_available_langs(self, service="yandex"):
        '''
        Returns the list of available languages
        '''
        return list(self.lang_to_code.keys())
        #return list(self.lang_to_code.keys())
    def get_detected_lang (self, word):
        '''
        Returns the language name of the word
        '''
        return self.code_to_lang[self.tr_impl.get_language_code(word)] 
    def __fixed_translations_list(self, translations: list):
        '''
        Auxiliary function to fix the list of translations for the word
        '''
        for i in range(len(translations)):
            translations[i] = translations[i].lower().strip().replace(chr(833), "")
        translations = list(set(translations))
        return translations 
    def get_translations (self, lang_from: str, lang_to: str, word: str) -> list:
        '''
        Returns saved translations for requested !!!word!!! or requests yandex cloud for translation if there are no available translations for this word
        and saves the history of request
        '''
        from_to_name = f"{lang_from}_to_{lang_to}"
        translations_ya_cl = [self.tr_impl.get_translation(word, hint_lang_codes=[lang_from], target_lang_code=lang_to)['text']]
        translations_dfs = list(self.dfs[from_to_name][self.dfs[from_to_name]['Word'] == word]['Translation'].values)
        translation_google = [GoogleTranslator(source=lang_from, target=lang_to).translate(word)]
        translations_pons = PonsTranslator(source=lang_from, target=lang_to).translate(word, return_all=True)
        translations = translations_ya_cl + translations_dfs + translations_pons + translation_google
        return self.__fixed_translations_list(translations)
    def get_translation (self, lang_from_code, lang_to_code, word, service="yandex"):
        '''
        Returns the tranlation of the requested word and saves the request to the history if the word is short enough
        '''
        if service == "yandex":
            translation = self.tr_impl.get_translation(word, hint_lang_codes=[lang_from_code], target_lang_code=lang_to_code)['text']
        else:
            translation = GoogleTranslator(source=lang_from_code, target=lang_to_code).translate(word)
        if len(translation) < self.save_to_histoty_limit or len(word) < self.save_to_histoty_limit:
            self.__add_history(word, translation)
        return translation
    def save_translation(self, lang_from, lang_to, word_from, word_to, hint=""):
        '''
        Saves the translation of the word into workbook and saves the workbook, also appends this to the local storage
        '''
        from_to_name = f"{lang_from}_to_{lang_to}"
        to_from_name = f"{lang_to}_to_{lang_from}"
        if from_to_name not in self.dfs.keys():
            self.__init_sheets(from_to_name, to_from_name)
        self.__add_translation(from_to_name, word_from, word_to, hint)
        self.__add_translation(to_from_name, word_to, word_from, hint)
        self.wb.save(self.settings['dict_file_name'])    
    def __add_translation(self, name, word_from, word_to, hint):
        self.wb[name].append([word_from, word_to, hint])
        self.dfs[name] = pd.concat([self.dfs[name], pd.DataFrame(
            [[word_from, word_to, hint]], columns=self.default_headers)])
    def __init_sheets(self, *args):
        '''
        Initializes the sheets in the workbook and saves the workbook, also appends this to the local storage
        '''
        for name in args:
            assert(type(name) == str)
            self.wb.create_sheet(name)
            self.wb[name].append(self.default_headers)
            self.wb.save(self.settings['dict_file_name'])
            self.dfs[name] = pd.DataFrame(columns=self.default_headers)
    def __init_settings_file(self):
        '''
        Initializing the pyeng_settings.json file.
        If file is not found showing the error message and raising exception.
        '''
        self.__enter_cwd()
        settings_file_path = "pyeng_settings.json" 
        if not os.path.exists(settings_file_path):
            showerror("PyEng", self.__no_settings_file_error_msg())
            raise Exception(self.__no_settings_file_error_msg())
        self.settings = json.loads(open(settings_file_path, "r", encoding='utf-8').read())
    def __init_dictionaries(self):
        '''
        Initializing dictionaries as dataframes and openpyxl workbook (settings dict should be already set) 
        '''
        self.dfs = {}
        self.settings["dict_file_name"] = self.settings["dict_file_name"]#.encode('utf-8').decode(sys.getfilesystemencoding())
        if not os.path.exists(self.settings["dict_file_name"]):
            showerror("PyEng", self.__no_dict_file_error_msg())
            self.settings["dict_file_name"] = tk.filedialog.askopenfilename(
                title="Select dictionary excel file", filetypes=[("EXCEL", "*.xlsx")]).encode('utf-8').decode('utf-8')
        self.wb = openpyxl.load_workbook(self.settings["dict_file_name"])
        if "History" not in self.wb.sheetnames:
            self.__init_sheets("History")
        for sheetname in self.wb.sheetnames:
            self.dfs[sheetname] = pd.read_excel(self.settings['dict_file_name'], sheet_name=sheetname, engine='openpyxl')
    def __add_history(self, word, translation):
        '''
        Adds the word and its translation to the history sheet
        '''
        self.wb["History"].append([word, translation])
        self.wb.save(self.settings['dict_file_name'])
    def __enter_cwd(self):
        '''
        Enteres the current working directory
        '''
        os.chdir(os.path.dirname(os.path.abspath(__file__)))
    def __no_dict_file_error_msg(self):
        '''
        Returns the error message if dictionary file is not found
        '''
        return f"Excel file with name {self.settings['dict_file_name']} not found"
    def __no_settings_file_error_msg(self):
        '''
        Returns the error message if settings file is not found
        '''
        return (f"In order to make application work you have to create settings file \n" +
                  f"in {os.path.dirname(os.path.abspath(__file__))} \n" +
                  "with name 'pyeng_settings.json'\n" +
                  "with the following structure: \n"+
                  "{\n" +
                  "\tlangs: [<lang1>, <lang2>],\n"
                  "\tdict_file_name: <dict_file_name>,\n"+
                  "\tapi_key: <api_key>\n"+
                  "\tfolder_id: <folder_id>\n"+
                  "}")
    