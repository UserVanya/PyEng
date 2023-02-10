import openpyxl
from tkinter.simpledialog import askstring
from tkinter.filedialog import askopenfilename
import tkinter as tk
from tkinter import ttk
from tkinter import scrolledtext
import pandas as pd
import os
import json
from pyeng_translator_impl import TranslatorImpl

'''
Translation application:
Application requires settings file in json format with the following structure:
{
    "api_key": <api_key>,
    "folder_id": <folder_id>,
    "dict_file_name": <dict_file_name>,
    "langs": [<lang1>, <lang2>, ...]
}
Application works as follows:
It contatins two text fields: one for input and one for output.
When user types in input field, application translates the word and shows the translation in output field.
'''

class TranslatorApp:
    def __init__(self, master):
        self.enter_cwd()
        settings_file_path = "pyeng_settings.json"
        if not os.path.exists(settings_file_path):
            tk.messagebox.showinfo("PyEng", "Please select settings file")
            settings_file_path = askopenfilename(title="Select settings file", filetypes=[("JSON", "*.json")], parent=master)        
        settings_dict = json.loads(open(settings_file_path, "r").read())
        self.impl = TranslatorImpl(settings_dict["api_key"],settings_dict["folder_id"], settings_dict['langs'])
        self.dict_wb_name = settings_dict['dict_file_name']
        self.init_excel_table(self.dict_wb_name)
        self.wb = openpyxl.load_workbook(self.dict_wb_name)
        self.eng_to_rus_df = pd.read_excel(self.dict_wb_name, sheet_name="eng_to_rus")
        self.rus_to_eng_df = pd.read_excel(self.dict_wb_name, sheet_name="rus_to_eng")
        self.master = master
        self.master.title("PyEng")
        self.master.geometry("600x600")
        self.master.resizable(True, True)
        self.master.configure(bg="#f0f0f0")
        master.grid_columnconfigure(0, weight = 1)
        master.grid_columnconfigure(1, weight = 1)
        master.grid_rowconfigure(0, weight = 5)
        master.grid_rowconfigure(1, weight = 1)
        master.grid_rowconfigure(2, weight = 1)
        self.create_widgets()
    def get_settings_dict (self, filename):
        return json.loads(open(filename, "r").read())
    def get_current_root_geometry(self) -> dict:
        
        return self.master.winfo_geometry()
    def init_excel_table(self, name):
        self.enter_cwd()
        if not os.path.exists(name):
            wb = openpyxl.Workbook()
            wb.remove_sheet(wb.active)
            if "eng_to_rus" not in wb.sheetnames:
              wb.create_sheet("eng_to_rus")
            if "rus_to_eng" not in wb.sheetnames:
              wb.create_sheet("rus_to_eng")
            wb.save(name)
    def enter_cwd(self):
        os.chdir(os.path.dirname(os.path.abspath(__file__)))
    def create_widgets(self):
        self.left_frame = tk.Frame(self.master, bg="#f0f0f0")
        self.left_frame.grid(row=0, column=0, sticky="nsew")
        self.left_frame.grid_columnconfigure(0, weight = 5)
        self.left_frame.grid_columnconfigure(1, weight = 1)
        self.left_frame.grid_rowconfigure(0, weight = 1)
        self.input_text = tk.Text(self.left_frame, bg="#f0f0f0", wrap="word")
        self.input_text.grid(row=0, column=0, sticky="nsew")
        self.input_scrollbar = tk.Scrollbar(self.left_frame, orient="vertical", width=1, command=self.input_text.yview)
        self.input_text["yscrollcommand"]=self.input_scrollbar.set
        self.input_scrollbar.grid(row=0, column=1, sticky="nsew")

        self.right_frame = tk.Frame(self.master, bg="#f0f0f0")
        self.right_frame.grid(row=0, column=1, sticky="nsew")
        self.right_frame.grid_columnconfigure(0, weight = 5)
        self.right_frame.grid_columnconfigure(1, weight = 1)
        self.right_frame.grid_rowconfigure(0, weight = 1)
        self.output_text = tk.Text(self.right_frame, bg="#f0f0f0", wrap="word", state="disabled")
        self.output_text.grid(row=0, column=0, sticky="nsew")
        self.output_scrollbar = tk.Scrollbar(self.right_frame, orient="vertical", width=1, command=self.output_text.yview)
        self.output_text["yscrollcommand"]=self.output_scrollbar.set
        self.output_scrollbar.grid(row=0, column=1, sticky="nsew")

        self.detected_lang_frame = tk.Frame(self.master, bg="#f0f0f0")
        self.detected_lang_frame.grid(row=1, column=0, columnspan = 2 ,sticky="nsew")
        self.detected_lang_var = tk.StringVar()
        self.detected_lang_label = tk.Label(self.detected_lang_frame, textvariable=self.detected_lang_var, bg="#f0f0f0")
        self.detected_lang_label.pack(fill="both", expand=True)
        
        self.translate_button = tk.Button(self.master, text="Translate", command=self.translate)
        self.translate_button.grid(row=2, column=0, columnspan = 2, sticky="nsew")
    def translate(self):
        text = self.input_text.get("1.0", "end-1c")
        self.detected_lang_var.set("Detected language: " + self.impl.get_language_code(text))
        self.output_text.config(state="normal")
        self.output_text.delete("1.0", "end")
        self.output_text.insert(tk.END, self.impl.translate(text))
        self.output_text.config(state="disabled")
if __name__ == "__main__":
    root = tk.Tk()
    app = TranslatorApp(root)
    root.mainloop()